import pandas as pd
from typing import List, Dict, Any
import io
from datetime import datetime
import os


class ExcelService:
    def __init__(self):
        pass
    
    def create_excel_from_invoices(self, invoice_data_list: List[Dict[str, Any]]) -> bytes:
        """Create Excel file with Invoice Summary and Line Items sheets"""
        
        # Prepare data for both sheets
        summary_data = []
        line_items_data = []
        
        for invoice_data in invoice_data_list:
            if 'invoice_summary' in invoice_data:
                summary_data.append(invoice_data['invoice_summary'])
            
            if 'line_items' in invoice_data and invoice_data['line_items']:
                for item in invoice_data['line_items']:
                    # Add invoice info to each line item
                    enhanced_item = item.copy()
                    if 'invoice_summary' in invoice_data:
                        enhanced_item['invoice_number'] = invoice_data['invoice_summary'].get('invoice_number', 'N/A')
                        enhanced_item['vendor_name'] = invoice_data['invoice_summary'].get('vendor_name', 'Unknown')
                        enhanced_item['invoice_date'] = invoice_data['invoice_summary'].get('invoice_date', 'N/A')
                    line_items_data.append(enhanced_item)
        
        # Create DataFrames
        summary_df = pd.DataFrame(summary_data) if summary_data else pd.DataFrame()
        line_items_df = pd.DataFrame(line_items_data) if line_items_data else pd.DataFrame()
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Invoice Summary Sheet
            if not summary_df.empty:
                # Reorder columns for better readability
                summary_columns = [
                    'filename', 'invoice_number', 'vendor_name', 'invoice_date', 
                    'due_date', 'subtotal', 'tax_amount', 'total_amount', 
                    'po_number', 'currency', 'payment_terms', 'vendor_address'
                ]
                
                # Only include columns that exist
                available_columns = [col for col in summary_columns if col in summary_df.columns]
                summary_df = summary_df[available_columns]
                
                summary_df.to_excel(writer, sheet_name='Invoice Summary', index=False)
                
                # Format the summary sheet
                worksheet = writer.sheets['Invoice Summary']
                self._format_summary_sheet(worksheet, summary_df)
            
            # Line Items Sheet
            if not line_items_df.empty:
                # Reorder columns for line items
                line_item_columns = [
                    'filename', 'invoice_number', 'vendor_name', 'invoice_date',
                    'item_description', 'category', 'quantity', 'unit_price', 
                    'line_total'
                ]
                
                # Only include columns that exist
                available_columns = [col for col in line_item_columns if col in line_items_df.columns]
                line_items_df = line_items_df[available_columns]
                
                line_items_df.to_excel(writer, sheet_name='Line Items', index=False)
                
                # Format the line items sheet
                worksheet = writer.sheets['Line Items']
                self._format_line_items_sheet(worksheet, line_items_df)
            
            # Add summary statistics sheet
            self._add_summary_stats_sheet(writer, summary_df, line_items_df)
        
        output.seek(0)
        return output.getvalue()
    
    def _format_summary_sheet(self, worksheet, df):
        """Format the Invoice Summary sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_line_items_sheet(self, worksheet, df):
        """Format the Line Items sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_summary_stats_sheet(self, writer, summary_df, line_items_df):
        """Add a summary statistics sheet"""
        stats_data = []
        
        if not summary_df.empty:
            stats_data.append(['Total Invoices Processed', len(summary_df)])
            
            if 'total_amount' in summary_df.columns:
                total_amount = summary_df['total_amount'].sum()
                stats_data.append(['Total Amount (All Invoices)', f"${total_amount:,.2f}"])
                
                avg_amount = summary_df['total_amount'].mean()
                stats_data.append(['Average Invoice Amount', f"${avg_amount:,.2f}"])
            
            if 'vendor_name' in summary_df.columns:
                unique_vendors = summary_df['vendor_name'].nunique()
                stats_data.append(['Unique Vendors', unique_vendors])
        
        if not line_items_df.empty:
            stats_data.append(['Total Line Items', len(line_items_df)])
            
            if 'category' in line_items_df.columns:
                top_categories = line_items_df['category'].value_counts().head(5)
                stats_data.append(['', ''])  # Empty row
                stats_data.append(['Top Categories:', ''])
                for category, count in top_categories.items():
                    stats_data.append([f"  {category}", count])
        
        # Processing timestamp
        stats_data.append(['', ''])
        stats_data.append(['Processed On', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        
        # Create stats DataFrame
        stats_df = pd.DataFrame(stats_data, columns=['Metric', 'Value'])
        stats_df.to_excel(writer, sheet_name='Summary Stats', index=False)
        
        # Format stats sheet
        if 'Summary Stats' in writer.sheets:
            worksheet = writer.sheets['Summary Stats']
            self._format_stats_sheet(worksheet)
    
    def _format_stats_sheet(self, worksheet):
        """Format the summary stats sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Make first column wider
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 20 